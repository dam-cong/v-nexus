"""Nạp ngân hàng câu hỏi từ docs/V-Nexus_Content_Workbook_1.xlsx vào hệ thống.

Đọc sheet QUESTION_BANK (chỉ lấy dòng đã "Đã duyệt") của workbook nội dung
Global Success lớp 3–4, sinh ra:
  - docs/data/question-bank.json — toàn bộ câu hỏi, khớp domain/knowledge_graph.py
  - docs/data/test-sets.json — 6 bộ đề khảo sát (chẩn đoán) chia theo khối lớp
    (G3/G4) x độ khó (easy/medium/hard), mỗi bộ tham chiếu question_id trong
    question-bank.json.

Chạy: python -m tools.import_workbook_questions
(cần `openpyxl`, xem requirements.txt)
"""
import json
import os
from collections import defaultdict

import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORKBOOK_PATH = os.path.join(BASE_DIR, "docs", "V-Nexus_Content_Workbook_1.xlsx")
DATA_DIR = os.path.join(BASE_DIR, "docs", "data")

DIFFICULTY_MAP = {"1": "easy", "2": "medium", "3": "hard"}
PURPOSE_MAP = {"CHAN_DOAN": "diagnostic", "LUYEN_TAP": "practice"}
DIFFICULTY_LABEL_VI = {"easy": "Dễ", "medium": "Trung bình", "hard": "Khó"}
DIFFICULTY_ORDER = ["easy", "medium", "hard"]


def _clean(value):
    """Chuẩn hoá ô Excel: None/rỗng/"-" đều coi là không có giá trị."""
    if value is None:
        return None
    text = str(value).strip()
    return None if text in ("", "-") else text


def load_skill_names():
    """node_id -> ten_ky_nang từ sheet NODE_INPUT."""
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = wb["NODE_INPUT"]
    names = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        node_id = row[0]
        if node_id:
            names[node_id] = row[2]
    return names


def load_approved_rows():
    """Danh sách dict cho mỗi dòng "Đã duyệt" trong sheet QUESTION_BANK."""
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = wb["QUESTION_BANK"]
    rows_iter = sheet.iter_rows(min_row=1, values_only=True)
    header = next(rows_iter)

    rows = []
    for values in rows_iter:
        record = dict(zip(header, values))
        if not record.get("node_id"):
            continue
        status = _clean(record.get("trang_thai (Nháp/Đã duyệt/Loại)"))
        if status != "Đã duyệt":
            continue
        rows.append(record)
    return rows


def build_question(record, index, skill_names):
    node_id = record["node_id"]
    do_kho = str(record["do_kho (1-3)"]).strip()
    loai = str(record["loai (CHAN_DOAN/LUYEN_TAP)"]).strip()
    image_desc = _clean(record.get("image_desc"))

    options = []
    for letter in ("A", "B", "C"):
        label = _clean(record.get(f"dap_an_{letter}"))
        if label is None:
            continue
        options.append({
            "option_id": letter.lower(),
            "label": label,
            "icon": None,
            "error_tag": _clean(record.get(f"loi_{letter}")),
            "feedback_vi": _clean(record.get(f"phan_hoi_{letter}")),
        })

    return {
        "question_id": f"gsq_{index:03d}",
        "type": "listen_icon_choice" if image_desc else "listen_choice",
        "instruction_label": "NGHE & CHỌN",
        "skill_id": node_id,
        "skill_name": skill_names.get(node_id, node_id),
        "difficulty": DIFFICULTY_MAP.get(do_kho, "medium"),
        "purpose": PURPOSE_MAP.get(loai, "diagnostic"),
        "prompt": {
            "text": None,
            "audio_url": None,
            "audio_transcript": str(record["cau_hoi_audio_script"]).strip(),
            "image_desc": image_desc,
        },
        "options": options,
        "correct_option_id": str(record["dap_an_dung"]).strip().lower(),
        "explanation": None,
    }


def build_question_bank(questions):
    return {
        "bank_id": "en-global-success-3-4-qb",
        "version": "1.0.0",
        "source": "docs/V-Nexus_Content_Workbook_1.xlsx (sheet QUESTION_BANK) - SGK Global Success lop 3-4",
        "description": (
            "Ngan hang cau hoi Tieng Anh tieu hoc (Lop 3-4) theo SGK Global Success. "
            "Moi cau gan skill_id (= node_id trong workbook) khop domain/knowledge_graph.py "
            "de dung lam evidence cho BKT."
        ),
        "total_questions": len(questions),
        "skills_covered": sorted({q["skill_id"] for q in questions}),
        "questions": questions,
    }


def build_test_sets(questions):
    """6 bộ đề khảo sát: khối lớp (G3/G4) x độ khó (easy/medium/hard).

    Chỉ lấy câu purpose="diagnostic" (CHAN_DOAN) — bộ đề khảo sát dùng để chẩn
    đoán, không lẫn câu luyện tập (LUYEN_TAP).
    """
    buckets = defaultdict(list)
    for q in questions:
        if q["purpose"] != "diagnostic":
            continue
        grade = "3" if q["skill_id"].startswith("G3") else "4"
        buckets[(grade, q["difficulty"])].append(q)

    test_sets = []
    for grade in ["3", "4"]:
        for difficulty in DIFFICULTY_ORDER:
            qs = buckets.get((grade, difficulty), [])
            if not qs:
                continue
            test_sets.append({
                "test_id": f"gs-g{grade}-{difficulty}",
                "title": f"Khảo sát Tiếng Anh Global Success — Lớp {grade} ({DIFFICULTY_LABEL_VI[difficulty]})",
                "grade": int(grade),
                "difficulty": difficulty,
                "mascot": {"name": "Medi Bee", "icon": "bee"},
                "steps": [
                    {"order": 1, "key": "choose_level", "label": "Chọn cấp độ"},
                    {"order": 2, "key": "take_test", "label": "Làm bài"},
                    {"order": 3, "key": "get_result", "label": "Nhận kết quả"},
                ],
                "skills_covered": sorted({q["skill_id"] for q in qs}),
                "question_ids": [q["question_id"] for q in qs],
                "scoring": {
                    "total_questions": len(qs),
                    "points_per_correct": 1,
                    "max_score": len(qs),
                },
            })
    return test_sets


def main():
    skill_names = load_skill_names()
    records = load_approved_rows()
    questions = [build_question(r, i + 1, skill_names) for i, r in enumerate(records)]

    bank = build_question_bank(questions)
    test_sets = build_test_sets(questions)

    bank_path = os.path.join(DATA_DIR, "question-bank.json")
    with open(bank_path, "w", encoding="utf-8") as f:
        json.dump(bank, f, ensure_ascii=False, indent=2)

    test_sets_path = os.path.join(DATA_DIR, "test-sets.json")
    with open(test_sets_path, "w", encoding="utf-8") as f:
        json.dump({
            "description": (
                "6 bo de khao sat (chan doan) theo khoi lop x do kho, sinh tu "
                "question-bank.json (chi lay cau purpose=diagnostic)."
            ),
            "source": "docs/data/question-bank.json (bank_id en-global-success-3-4-qb)",
            "test_sets": test_sets,
        }, f, ensure_ascii=False, indent=2)

    print(f"Da ghi {len(questions)} cau hoi ({len(bank['skills_covered'])} skill) vao {bank_path}")
    print(f"Da tao {len(test_sets)} bo de khao sat vao {test_sets_path}")


if __name__ == "__main__":
    main()
